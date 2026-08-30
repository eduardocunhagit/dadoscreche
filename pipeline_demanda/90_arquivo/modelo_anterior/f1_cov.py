from pathlib import Path
import math

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import sparse
from sklearn.linear_model import PoissonRegressor
from sklearn.preprocessing import OneHotEncoder

from f1_est import (
    GROUPS,
    build_panel,
    load_births,
    load_unique_child_years,
)


# Reproducibility: estimation is deterministic, but the project seed is fixed by convention.
np.random.seed(42)

MODEL_DIR = Path(__file__).resolve().parent
DATA_DIR = MODEL_DIR / "data_ext"
RESULTS_DIR = MODEL_DIR / "results"
FIGURES_DIR = MODEL_DIR / "figures"

CAGED_PATH = DATA_DIR / "caged_2025.xlsx"
RAIS_FILES = {
    "rais_2021.xlsx": {"years": [2020, 2021], "kind": "comparison"},
    "rais_2022.xlsx": {"years": [2022], "kind": "single"},
    "rais_2022_2023.xlsx": {"years": [2022, 2023], "kind": "comparison"},
    "rais_2023_2024.xlsx": {"years": [2023, 2024], "kind": "comparison"},
}


def find_rio_row(sheet):
    """Return the municipality row without loading the full national sheet into memory."""
    for row in sheet.iter_rows(values_only=True):
        first_cells = [str(value).strip().lower() for value in row[:6] if value is not None]
        if any("rio de janeiro" in value for value in first_cells):
            return row
    raise ValueError(f"Rio de Janeiro was not found in sheet {sheet.title!r}.")


def load_caged_rio():
    """Extract the adjusted monthly Novo Caged series for the municipality of Rio."""
    workbook = load_workbook(CAGED_PATH, read_only=True, data_only=True)
    sheet = workbook["Tabela 8.1"]
    month_headers = next(sheet.iter_rows(min_row=5, max_row=5, values_only=True))
    measure_headers = next(sheet.iter_rows(min_row=6, max_row=6, values_only=True))
    rio_row = find_rio_row(sheet)

    records = []
    current_month = None
    for position, (month_label, measure_label) in enumerate(
        zip(month_headers, measure_headers), start=1
    ):
        if month_label is not None and "/" in str(month_label):
            current_month = str(month_label).strip()
        if measure_label == "Estoque" and current_month is not None:
            stock_index = position - 1
            records.append(
                {
                    "mes_rotulo": current_month,
                    "estoque": rio_row[stock_index],
                    "admissoes": rio_row[stock_index + 1],
                    "desligamentos": rio_row[stock_index + 2],
                    "saldo": rio_row[stock_index + 3],
                }
            )

    months_pt = {
        "Janeiro": 1,
        "Fevereiro": 2,
        "Março": 3,
        "Abril": 4,
        "Maio": 5,
        "Junho": 6,
        "Julho": 7,
        "Agosto": 8,
        "Setembro": 9,
        "Outubro": 10,
        "Novembro": 11,
        "Dezembro": 12,
    }
    monthly = pd.DataFrame(records)
    monthly[["mes_nome", "ano"]] = monthly["mes_rotulo"].str.split("/", expand=True)
    monthly["ano"] = monthly["ano"].astype(int)
    monthly["mes"] = monthly["mes_nome"].map(months_pt)
    monthly["data"] = pd.to_datetime(
        {"year": monthly["ano"], "month": monthly["mes"], "day": 1}
    )
    numeric_columns = ["estoque", "admissoes", "desligamentos", "saldo"]
    monthly[numeric_columns] = monthly[numeric_columns].apply(pd.to_numeric, errors="coerce")

    annual_rows = []
    for year, year_data in monthly.groupby("ano", sort=True):
        year_data = year_data.sort_values("mes")
        first_row = year_data.iloc[0]
        last_row = year_data.iloc[-1]
        reconstructed_start_stock = first_row["estoque"] - first_row["saldo"]
        annual_balance = year_data["saldo"].sum()
        annual_rows.append(
            {
                "ano_indicador": int(year),
                "estoque_inicio_reconstruido": reconstructed_start_stock,
                "estoque_dezembro": last_row["estoque"],
                "admissoes_ano": year_data["admissoes"].sum(),
                "desligamentos_ano": year_data["desligamentos"].sum(),
                "saldo_ano": annual_balance,
                "taxa_saldo_estoque_inicio": annual_balance / reconstructed_start_stock,
            }
        )
    annual = pd.DataFrame(annual_rows)
    return monthly, annual


def load_rais_vintages():
    """Extract Rio employment stocks from each official RAIS workbook vintage."""
    records = []
    for file_name, specification in RAIS_FILES.items():
        path = DATA_DIR / file_name
        workbook = load_workbook(path, read_only=True, data_only=True)
        rio_row = find_rio_row(workbook["TABELA 4"])
        non_null_values = [value for value in rio_row if value is not None]
        if specification["kind"] == "single":
            values = [non_null_values[-2]]
        else:
            values = [non_null_values[-4], non_null_values[-3]]
        for year, value in zip(specification["years"], values):
            records.append(
                {
                    "arquivo_vintage": file_name,
                    "ano_rais": year,
                    "estoque_formal": int(value),
                }
            )
    vintages = pd.DataFrame(records)
    vintages["revisao_vs_primeira_pct"] = (
        vintages.groupby("ano_rais")["estoque_formal"].transform(
            lambda values: 100 * (values / values.iloc[0] - 1)
        )
    )
    return vintages


def add_caged_covariates(panel, annual_caged):
    """Attach only information from the calendar year preceding the demand year."""
    lagged = annual_caged.rename(
        columns={
            "ano_indicador": "ano_caged_lag1",
            "taxa_saldo_estoque_inicio": "caged_taxa_lag1",
            "estoque_dezembro": "caged_estoque_dez_lag1",
        }
    ).copy()
    lagged["ano"] = lagged["ano_caged_lag1"] + 1
    keep = ["ano", "ano_caged_lag1", "caged_taxa_lag1", "caged_estoque_dez_lag1"]
    return panel.merge(lagged[keep], how="left", on="ano")


def make_design_matrix(frame, specification, encoder=None, scaler=None, fit=False):
    """Build each predeclared specification using only training-sample scaling."""
    categorical = frame[["territorio", "grupamento_modelo"]]
    if fit:
        try:
            encoder = OneHotEncoder(handle_unknown="ignore", sparse_output=True)
        except TypeError:
            encoder = OneHotEncoder(handle_unknown="ignore", sparse=True)
        categorical_matrix = encoder.fit_transform(categorical)
    else:
        categorical_matrix = encoder.transform(categorical)

    matrices = [categorical_matrix]
    if specification["include_trends"]:
        trend_columns = [column for column in frame.columns if column.startswith("tend_")]
        matrices.append(sparse.csr_matrix(frame[trend_columns].to_numpy(dtype=float)))

    if specification["caged_mode"] != "none":
        raw_feature = frame[["caged_taxa_lag1"]].to_numpy(dtype=float)
        if fit:
            feature_mean = float(np.nanmean(raw_feature))
            feature_std = float(np.nanstd(raw_feature))
            scaler = {"mean": feature_mean, "std": feature_std if feature_std > 0 else 1.0}
        standardized = (raw_feature - scaler["mean"]) / scaler["std"]
        if specification["caged_mode"] == "common":
            matrices.append(sparse.csr_matrix(standardized))
        elif specification["caged_mode"] == "group_interactions":
            interactions = []
            for group in GROUPS:
                group_indicator = (frame["grupamento_modelo"] == group).to_numpy(dtype=float)
                interactions.append(standardized[:, 0] * group_indicator)
            matrices.append(sparse.csr_matrix(np.column_stack(interactions)))
        else:
            raise ValueError(f"Unknown Caged mode: {specification['caged_mode']}")

    return sparse.hstack(matrices, format="csr"), encoder, scaler


def fit_model(train_frame, specification):
    estimation_sample = train_frame[
        (train_frame["universo_elegivel_proxy"] > 0)
        & train_frame["caged_taxa_lag1"].notna()
    ].copy()
    design, encoder, scaler = make_design_matrix(
        estimation_sample, specification, fit=True
    )
    exposure = estimation_sample["universo_elegivel_proxy"].to_numpy(dtype=float)
    rate = estimation_sample["inscritos"].to_numpy(dtype=float) / exposure
    model = PoissonRegressor(alpha=1e-4, fit_intercept=True, max_iter=3000, tol=1e-9)
    model.fit(design, rate, sample_weight=exposure)
    return model, encoder, scaler, estimation_sample


def predict_model(model, encoder, scaler, frame, specification):
    prediction_sample = frame[
        (frame["universo_elegivel_proxy"] > 0) & frame["caged_taxa_lag1"].notna()
    ].copy()
    design, _, _ = make_design_matrix(
        prediction_sample,
        specification,
        encoder=encoder,
        scaler=scaler,
        fit=False,
    )
    prediction_sample["previsto"] = (
        model.predict(design)
        * prediction_sample["universo_elegivel_proxy"].to_numpy(dtype=float)
    )
    return prediction_sample


def calculate_metrics(predictions, model_name, test_year):
    actual = predictions["inscritos"].to_numpy(dtype=float)
    predicted = predictions["previsto"].to_numpy(dtype=float)
    errors = predicted - actual
    observed_total = actual.sum()
    return {
        "ano_teste": test_year,
        "modelo": model_name,
        "n_celulas": len(predictions),
        "observado_total": observed_total,
        "previsto_total": predicted.sum(),
        "erro_total": errors.sum(),
        "mae": np.mean(np.abs(errors)),
        "rmse": math.sqrt(np.mean(errors**2)),
        "wape": np.sum(np.abs(errors)) / observed_total,
        "vies_relativo": errors.sum() / observed_total,
    }


def add_persistence(panel, predictions, test_year, estimation_sample):
    previous = panel[panel["ano"] == test_year - 1][
        ["territorio", "grupamento_modelo", "inscritos", "universo_elegivel_proxy"]
    ].copy()
    previous["taxa_anterior"] = previous["inscritos"] / previous[
        "universo_elegivel_proxy"
    ].replace(0, np.nan)
    predictions = predictions.merge(
        previous[["territorio", "grupamento_modelo", "taxa_anterior"]],
        how="left",
        on=["territorio", "grupamento_modelo"],
    )
    fallback_rate = (
        estimation_sample.groupby("grupamento_modelo")["inscritos"].sum()
        / estimation_sample.groupby("grupamento_modelo")["universo_elegivel_proxy"].sum()
    )
    predictions["taxa_anterior"] = predictions["taxa_anterior"].fillna(
        predictions["grupamento_modelo"].map(fallback_rate)
    )
    predictions["prev_persistencia"] = (
        predictions["taxa_anterior"] * predictions["universo_elegivel_proxy"]
    )
    return predictions


def run_oos(panel, specification, test_year):
    train = panel[(panel["ano"] >= 2021) & (panel["ano"] < test_year)].copy()
    test = panel[panel["ano"] == test_year].copy()
    model, encoder, scaler, estimation_sample = fit_model(train, specification)
    predictions = predict_model(model, encoder, scaler, test, specification)
    predictions = add_persistence(panel, predictions, test_year, estimation_sample)
    predictions["ano_teste"] = test_year
    predictions["modelo"] = specification["name"]
    metrics = calculate_metrics(predictions, specification["name"], test_year)
    return predictions, metrics


def run_forecast_2026(panel, specification):
    train = panel[(panel["ano"] >= 2021) & (panel["ano"] <= 2025)].copy()
    test = panel[panel["ano"] == 2026].copy()
    model, encoder, scaler, estimation_sample = fit_model(train, specification)
    forecast = predict_model(model, encoder, scaler, test, specification)
    forecast = add_persistence(panel, forecast, 2026, estimation_sample)
    forecast["modelo"] = specification["name"]
    return forecast


def persistence_metrics(predictions, test_year):
    persistence = predictions.copy()
    persistence["previsto"] = persistence["prev_persistencia"]
    return calculate_metrics(persistence, "Persistencia taxa t-1", test_year)


def save_oos_figure(metrics):
    plot_data = metrics[metrics["modelo"] != "Persistencia taxa t-1"].copy()
    plot_data["wape_pct"] = 100 * plot_data["wape"]
    models = list(plot_data["modelo"].drop_duplicates())
    years = sorted(plot_data["ano_teste"].unique())
    x = np.arange(len(years))
    width = 0.8 / len(models)
    colors = ["#17324D", "#2C7FB8", "#41AB5D", "#F28E2B"]
    fig, ax = plt.subplots(figsize=(10.5, 5.3))
    for index, model_name in enumerate(models):
        part = plot_data[plot_data["modelo"] == model_name].set_index("ano_teste")
        values = [part.loc[year, "wape_pct"] for year in years]
        positions = x - 0.4 + width / 2 + index * width
        ax.bar(positions, values, width, label=model_name, color=colors[index])
    ax.set_xticks(x, [str(year) for year in years])
    ax.set_ylabel("WAPE (%)")
    ax.set_title("Teste OOS: covariada municipal do Novo Caged")
    ax.grid(axis="y", alpha=0.2)
    ax.legend(frameon=False, fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "f1_cov_oos.png", dpi=180)
    plt.close(fig)


def main():
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    monthly_caged, annual_caged = load_caged_rio()
    rais_vintages = load_rais_vintages()
    _, child_years = load_unique_child_years()
    neighborhood_dimension, births_long = load_births()
    panel, _ = build_panel(child_years, neighborhood_dimension, births_long)
    panel = add_caged_covariates(panel, annual_caged)

    specifications = [
        {"name": "PPML-FE base replicado", "include_trends": True, "caged_mode": "none"},
        {"name": "PPML-FE + Caged comum", "include_trends": True, "caged_mode": "common"},
        {
            "name": "PPML-FE + Caged x grupo",
            "include_trends": True,
            "caged_mode": "group_interactions",
        },
        {
            "name": "PPML-FE Caged x grupo sem tendencia",
            "include_trends": False,
            "caged_mode": "group_interactions",
        },
    ]

    all_predictions = []
    all_metrics = []
    for test_year in [2024, 2025]:
        persistence_added = False
        for specification in specifications:
            predictions, metrics = run_oos(panel, specification, test_year)
            all_predictions.append(predictions)
            all_metrics.append(metrics)
            if not persistence_added:
                all_metrics.append(persistence_metrics(predictions, test_year))
                persistence_added = True

    predictions = pd.concat(all_predictions, ignore_index=True)
    metrics = pd.DataFrame(all_metrics)
    baseline = metrics[metrics["modelo"] == "PPML-FE base replicado"][
        ["ano_teste", "wape"]
    ].rename(columns={"wape": "wape_base"})
    metrics = metrics.merge(baseline, how="left", on="ano_teste")
    metrics["delta_wape_pp"] = 100 * (metrics["wape"] - metrics["wape_base"])

    pooled_metrics = (
        metrics.assign(
            erro_abs_total=lambda frame: frame["wape"] * frame["observado_total"],
            vies_total=lambda frame: frame["vies_relativo"] * frame["observado_total"],
        )
        .groupby("modelo", as_index=False)
        .agg(
            observado_total=("observado_total", "sum"),
            previsto_total=("previsto_total", "sum"),
            erro_abs_total=("erro_abs_total", "sum"),
            vies_total=("vies_total", "sum"),
        )
    )
    pooled_metrics["wape_pooled"] = (
        pooled_metrics["erro_abs_total"] / pooled_metrics["observado_total"]
    )
    pooled_metrics["vies_relativo_pooled"] = (
        pooled_metrics["vies_total"] / pooled_metrics["observado_total"]
    )

    forecast_2026 = pd.concat(
        [run_forecast_2026(panel, specification) for specification in specifications],
        ignore_index=True,
    )
    forecast_summary = (
        forecast_2026.groupby("modelo", as_index=False)
        .agg(
            previsao_2026=("previsto", "sum"),
            persistencia_2026=("prev_persistencia", "sum"),
        )
    )

    availability = pd.DataFrame(
        [
            {
                "fonte": "Novo Caged",
                "periodo_localizado": "2020-2025 mensal",
                "menor_geografia_publica": "municipio",
                "variacao_entre_bairros": "nao",
                "uso_no_teste": "taxa anual t-1 comum a todas as creches",
                "veredito": "testar; nao interpretar como emprego residencial local",
            },
            {
                "fonte": "RAIS",
                "periodo_localizado": "2020-2024 anual",
                "menor_geografia_publica": "municipio",
                "variacao_entre_bairros": "nao",
                "uso_no_teste": "auditoria de estoque e revisoes",
                "veredito": "nao incluir junto ao Caged: atraso, quebra e pouca dimensao temporal",
            },
            {
                "fonte": "Habite-se Data.Rio",
                "periodo_localizado": "2009-2015; bairro/RA/AP",
                "menor_geografia_publica": "bairro",
                "variacao_entre_bairros": "sim, mas fora do painel",
                "uso_no_teste": "nao estimado",
                "veredito": "solicitar extracao 2020-2025 do licenciamento municipal",
            },
        ]
    )

    best_caged = metrics[
        metrics["modelo"].str.contains("Caged", case=False, na=False)
    ].sort_values(["ano_teste", "wape"])
    decisions = []
    for year in [2024, 2025]:
        best_year = best_caged[best_caged["ano_teste"] == year].iloc[0]
        decisions.append(
            {
                "ano_teste": year,
                "melhor_especificacao_caged": best_year["modelo"],
                "wape_pct": 100 * best_year["wape"],
                "delta_wape_pp_vs_base": best_year["delta_wape_pp"],
                "aceitar_caged": "sim" if best_year["delta_wape_pp"] < 0 else "nao",
            }
        )
    decisions = pd.DataFrame(decisions)
    decisions["decisao_final"] = (
        "nao incorporar ao modelo principal: a melhora de 2024 nao se repetiu em 2025"
    )

    monthly_caged.to_csv(RESULTS_DIR / "f1_cov_caged_mensal.csv", index=False, encoding="utf-8-sig")
    annual_caged.to_csv(RESULTS_DIR / "f1_cov_caged_anual.csv", index=False, encoding="utf-8-sig")
    rais_vintages.to_csv(RESULTS_DIR / "f1_cov_rais_vintages.csv", index=False, encoding="utf-8-sig")
    availability.to_csv(RESULTS_DIR / "f1_cov_disponibilidade.csv", index=False, encoding="utf-8-sig")
    panel.to_csv(RESULTS_DIR / "f1_cov_panel.csv", index=False, encoding="utf-8-sig")
    predictions.to_csv(RESULTS_DIR / "f1_cov_oos_pred.csv", index=False, encoding="utf-8-sig")
    metrics.to_csv(RESULTS_DIR / "f1_cov_oos.csv", index=False, encoding="utf-8-sig")
    pooled_metrics.to_csv(RESULTS_DIR / "f1_cov_oos_pooled.csv", index=False, encoding="utf-8-sig")
    decisions.to_csv(RESULTS_DIR / "f1_cov_decisao.csv", index=False, encoding="utf-8-sig")
    forecast_2026.to_csv(RESULTS_DIR / "f1_cov_prev_2026.csv", index=False, encoding="utf-8-sig")
    forecast_summary.to_csv(RESULTS_DIR / "f1_cov_resumo_2026.csv", index=False, encoding="utf-8-sig")

    annual_caged.round(5).to_latex(RESULTS_DIR / "f1_cov_caged_anual.tex", index=False)
    rais_vintages.round(4).to_latex(RESULTS_DIR / "f1_cov_rais_vintages.tex", index=False)
    availability.to_latex(RESULTS_DIR / "f1_cov_disponibilidade.tex", index=False)
    metrics.round(5).to_latex(RESULTS_DIR / "f1_cov_oos.tex", index=False)
    pooled_metrics.round(5).to_latex(RESULTS_DIR / "f1_cov_oos_pooled.tex", index=False)
    decisions.round(4).to_latex(RESULTS_DIR / "f1_cov_decisao.tex", index=False)
    forecast_summary.round(2).to_latex(RESULTS_DIR / "f1_cov_resumo_2026.tex", index=False)
    save_oos_figure(metrics)

    print("NOVO CAGED - ANUAL")
    print(annual_caged.round(5).to_string(index=False))
    print("\nRAIS - VINTAGES")
    print(rais_vintages.round(4).to_string(index=False))
    print("\nOOS")
    print(metrics.round(5).to_string(index=False))
    print("\nDECISAO")
    print(decisions.round(4).to_string(index=False))
    print("\nOOS COMBINADO")
    print(pooled_metrics.round(5).to_string(index=False))
    print("\nSENSIBILIDADE 2026")
    print(forecast_summary.round(2).to_string(index=False))


if __name__ == "__main__":
    main()
